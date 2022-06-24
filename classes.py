from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.drawing.image import Image


class StocksReader:
    def __init__(self, path: str = ''):
        self.path = path
        self.data = []

    def file_process(self, stock: str):
        with open(f'{self.path}{stock}.txt', 'r') as stock_file:
            lines = stock_file.readlines()
            self.data = [line.replace('/n', '').split(';') for line in lines]


class ChartSeriesProperties:
    def __init__(self, width: int, solid_fill_color: str):
        self.width = width
        self.solid_fill_color = solid_fill_color


class SpreadsheetManager:
    def __init__(self):
        self.workbook = Workbook()
        self.active_spreadsheet = None

    def add_spreadsheet(self, title: str = ''):
        new_spreadsheet = self.workbook.create_sheet(title)
        self.workbook.active = new_spreadsheet
        self.active_spreadsheet = new_spreadsheet

        return new_spreadsheet

    def add_line(self, data: list):
        self.active_spreadsheet.append(data)

    def update_cell(self, cell: str, data):
        self.active_spreadsheet[cell] = data

    def merge_spreadsheet_cells(self, start_cell: str, end_cell: str):
        self.active_spreadsheet.merge_cells(f'{start_cell}:{end_cell}')

    def apply_style(self, cell: str, styles: list):
        for style in styles:
            setattr(self.active_spreadsheet[cell], style[0], style[1])

    def add_line_chart(self, cell: str, width: float, height: float,
                       title: str, x_axis_title: str, y_axis_title: str,
                       x_axis_reference: Reference, y_axis_reference: Reference,
                       chart_properties: list):
        chart = LineChart()
        chart.width = width
        chart.height = height
        chart.title = title
        chart.x_axis.title = x_axis_title
        chart.y_axis.title = y_axis_title

        chart.add_data(x_axis_reference)
        chart.set_categories(y_axis_reference)

        for serie, chart_prop in zip(chart.series, chart_properties):
            serie.graphicalProperties.line.width = chart_prop.width
            serie.graphicalProperties.line.solidFill = chart_prop.solid_fill_color

        self.active_spreadsheet.add_chart(chart, cell)

    def add_spreadsheet_image(self, cell: str, image_path: str):
        image = Image(image_path)
        self.active_spreadsheet.add_image(image, cell)

    def save_file(self, file_path: str):
        self.workbook.save(file_path)
